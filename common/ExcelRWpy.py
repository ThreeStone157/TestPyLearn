#coding: utf-8

import openpyxl

class ExcelRW:
    file_path = ""
    excel_workbook = ""
    sheet_list = ""
    def __init__(self, file_path):
        self.file_path = file_path
        self.excel_workbook = openpyxl.load_workbook(self.file_path)
        self.sheet_list = self.excel_workbook.sheetnames

    #获取指定sheet的最大行、列数
    def read_excel(self, sheetName = None):
        if sheetName in self.sheet_list:
            mySheet = self.excel_workbook.get_sheet_by_name(sheetName)
            max_row = mySheet.max_row
            max_column = mySheet.max_column
            print("最大行数为%s,最大列数为%s"%(max_row, max_column))
            return max_row, max_column
        else:
            return 0, 0

    # 获取单元格内容：
    # 1.如果传了指定的Sheet就获取指定的sheet内容返回
    # 2.如果有传行列，那么获取指定的行列数据返回。未传就直接读取对应的sheet内所有数据返回
    def read_data(self, sheetName = None, row = None, col = None):
        if sheetName == None or sheetName == "":
            for sheet in self.sheet_list:
                cases = self.read_sheet(sheet, row, col)
                print("--------")
        else:
            cases =  self.read_sheet(sheetName, row, col)
            print("用例:%s"%cases)
            return cases

    #读取单个sheet数据,除指定sheet名称必填，其他行列选填
    #只填写了行，就获取当前行所有的数据
    #只填写了列，就获取当前列所有的数据
    def read_sheet(self, sheetName, row = None, col = None):
        cases = {}
        case = []
        sheetObj = self.excel_workbook.get_sheet_by_name(sheetName)
        if row == None:
            if col == None:
                max_row, max_column = self.read_excel(sheetName)
                i = 2
                while (i <= max_row):
                    j = 1
                    while j <= max_column:
                        data = sheetObj.cell(i, j).value
                        j += 1
                        # print(1111)
                        case.append(data)
                    cases.setdefault(i, case)
                    i += 1

            else:
                max_row, max_column = self.read_excel(sheetName)
                i = 2
                while (i <= max_row):
                    data = sheetObj.cell(i, col).value
                    # print(aa)
                    cases.append(data)
                    i += 1
        elif col == None:
                max_row, max_column = self.read_excel(sheetName)
                j = 1
                while j <= max_column:
                    data = sheetObj.cell(row, j).value
                    j += 1
                    cases.append(data)
                return cases
        else:
            aa = sheetObj.cell(row, col).value
            # print(aa)
        return cases

if __name__ == "__main__":
    execlRW = ExcelRW("F:\工作文件\接口自动化Demo.xlsx")
    # row, col = execlRW.read_excel("")
    cases = execlRW.read_data("勋章")
    for keys in cases.keys():
        case = cases.get(keys)
        for value in case:
            print(value)

    # execlRW.read_sheet("勋章")
