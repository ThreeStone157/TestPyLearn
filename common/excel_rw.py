# coding: utf-8
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Side, Border, colors


# 封装的读取excel数据功能类
class ExcelRW:
    file_path = ""
    excel_workbook = ""
    sheet_list = ""

    # 初始化的时候就开始获取文件路径、文件的所有sheet
    def __init__(self, file_path):
        self.file_path = file_path
        self.excel_workbook = openpyxl.load_workbook(self.file_path)
        self.sheet_list = self.excel_workbook.sheetnames

    # 获取指定sheet的最大行、列数
    def read_excel(self, sheetName=None):
        if sheetName in self.sheet_list:
            # mySheet = self.excel_workbook.get_sheet_by_name(sheetName)
            my_sheet = self.excel_workbook[sheetName]
            max_rows = my_sheet.max_row
            max_columns = my_sheet.max_column
            # print("%s最大行数为%s,最大列数为%s"%(sheetName, max_rows, max_columns))
            print("%s" % sheetName)
            return max_rows, max_columns
        else:
            return 0, 0

    # 获取所有的sheet
    def get_sheets(self):
        # sheetss = self.excel_workbook.sheetnames
        # print(sheetss)
        return self.sheet_list

    # 获取单元格内容：
    # 1.如果传了指定的Sheet就获取指定的sheet内容返回
    # 2.如果有传行列，那么获取指定的行列数据返回。未传就直接读取对应的sheet内所有数据返回
    def read_data(self, sheetName=None, row=None, col=None):
        if sheetName is None or sheetName == "":
            for sheet_name in self.sheet_list:
                cases = self.read_sheet(sheet_name, row, col)
                print("--------")
        else:
            cases = self.read_sheet(sheetName, row, col)
            print("用例:%s" % cases)
        return cases

    # 读取单个sheet数据,除指定sheet名称必填，其他行列选填
    # 只填写了行，就获取当前行所有的数据
    # 只填写了列，就获取当前列所有的数据
    def read_sheet(self, sheetName, row = None, col = None):
        cases = {}
        case = []
        sheet_obj = self.excel_workbook.get_sheet_by_name(sheetName)
        if row is None:
            if col is None:
                max_rows, max_columns = self.read_excel(sheetName)
                i = 2
                while i <= max_rows:
                    j = 1
                    while j <= max_columns:
                        data = sheet_obj.cell(i, j).value
                        j += 1
                        # print(1111)
                        case.append(data)
                    cases.setdefault(i, case)
                    i += 1
            else:
                max_rows, max_columns = self.read_excel(sheetName)
                i = 2
                while i <= max_rows:
                    data = sheet_obj.cell(i, col).value
                    # print(aa)
                    cases.append(data)
                    i += 1
        elif col is None:
            max_rows, max_columns = self.read_excel(sheetName)
            j = 1
            while j <= max_columns:
                data = sheet_obj.cell(row, j).value
                j += 1
                cases.append(data)
            return cases
        else:
            aa = sheet_obj.cell(row, col).value
            # print(aa)
            return aa
        return cases


if __name__ == "__main__":
    execlRW = ExcelRW(r"E:\Work\和平精英\和平电视台用例整理_20220615.xlsx")
    # row, col = execlRW.read_excel()
    # cases = execlRW.read_data()
    # for keys in cases.keys():
    #     case = cases.get(keys)
    #     print(case)
    sums = 0
    sheets = execlRW.get_sheets()
    for sheet in sheets:
        max_row, max_column = execlRW.read_excel(sheet)
        sums += max_row-1
    print("总用例数为:%s" % sums)
    # execlRW.set_Format()
    for i in range(1, 50):
        print(i)



