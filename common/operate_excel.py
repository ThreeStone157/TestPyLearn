import openpyxl

class operate_excel:

    file_path = ""
    excel_workbook = ""
    sheet_list = ""

    def __init__(self, file_path):
        self.file_path = file_path
        self.excel_workbook = openpyxl.load_workbook(self.file_path)
        self.sheet_list = self.excel_workbook.sheetnames

    #获取指定sheet的最大行、列数
    def read_max(self, sheetName):
        mySheet = self.excel_workbook.get_sheet_by_name(sheetName)
        max_row = mySheet.max_row
        max_column = mySheet.max_column
        # print("最大行数为%s,最大列数为%s"%(max_row, max_column))
        return max_row, max_column

    #获取一列中的所有 行数据
    def read_row(self, sheetName, col):
        datas = []
        max_row, max_column = self.read_max(sheetName)
        i = 2
        while(i <= max_row):
            data = self.read_data(sheetName, i, col)
            i+=1
            datas.append(data)
        return datas

    # 获取一行中的所有 列数据
    def read_col(self, sheetName, row, platid = None):
        datas = []
        dicts = {}
        max_row, max_column = self.read_max(sheetName)
        j = 1
        if platid == None:
            while (j <= max_column):
                data = self.read_data(sheetName, row, j)
                j+=1
                datas.append(data)
            return datas
        else:
            while (j <= max_column):
                data = self.read_data(sheetName, row, j)
                key = self.read_data(sheetName, 1, j)
                dicts.setdefault(key, data)
                j+=1
            datas.append(dicts)
            return datas

    #获取sheet中指定的单元格数据
    def read_data(self, sheetName, row, col):
        mySheet = self.excel_workbook.get_sheet_by_name(sheetName)
        return mySheet.cell(row, col).value

    #读取指定sheet下数据细化处理
    def read_sheet(self, sheetName):
        cases = []
        max_row, max_column = self.read_max(sheetName)
        i = 2
        while i <= max_row:
            case = self.read_col(sheetName, i, 1)
            cases.extend(case)
            i+=1
        return cases

    # 获取excel文件的用例
    # 1.如果传了指定的Sheet就获取指定的sheet内容返回
    # 2.如果有传行列，那么获取指定的行列数据返回。未传就直接读取对应的sheet内所有数据返回
    def read_datas(self, sheetName = None):
        cases_data = []
        if sheetName == None or sheetName == "":
            for sheet in self.sheet_list:
                cases = self.read_sheet(sheet)
                cases_data.extend(cases)
        else:
            cases =  self.read_sheet(sheetName)
            cases_data.extend(cases)
        return cases_data

if __name__ == "__main__":
    execlRW = operate_excel("F:\工作文件\接口自动化Demo.xlsx")
    # row, col = execlRW.read_excel("")
    cases_data = execlRW.read_datas()
    for case in cases_data:
        print(case)


